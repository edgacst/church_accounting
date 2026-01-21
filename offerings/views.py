from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from django.shortcuts import render, get_object_or_404
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from django.db import models
from members.models import ChurchMember
from offerings.models import Offering, TaxCertificateIssueLog
from datetime import date
import io
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib import colors
from django.http import HttpResponse
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

@login_required
def tax_certificate_list(request):
    """증명서 발급 대상 교인 목록"""
    # 마이페이지에서 온 경우 본인만, 관리자 메뉴에서 온 경우 모든 교인
    # URL이나 특정 파라미터로 구분하는 대신, 항상 본인 교인 정보를 우선 표시
    
    # 먼저 로그인한 사용자의 교인 정보를 찾음
    my_member = ChurchMember.objects.filter(
        models.Q(email=request.user.email) | 
        models.Q(korean_name=request.user.username)
    ).first()
    
    # 본인 교인 정보가 있으면 본인만, 없으면 관리자용으로 동의한 모든 교인
    if my_member:
        members = [my_member]
    elif request.user.is_staff:
        members = ChurchMember.objects.filter(tax_issuance_consent=True)
    else:
        members = []
    
    current_year = date.today().year
    
    # 각 교인별 헌금 합계 계산
    member_data = []
    for member in members:
        yearly_offerings = Offering.objects.filter(
            member=member,
            offering_date__year=current_year,
            is_confirmed=True
        )
        total = sum([o.amount for o in yearly_offerings])
        member_data.append({
            'member': member,
            'total_amount': total,
            'offering_count': yearly_offerings.count()
        })
    
    return render(request, 'offerings/tax_certificate_list.html', {
        'members': member_data,
        'current_year': current_year
    })

@login_required
def yearly_tax_certificate(request, member_id, year=None):
    """HTML로 출력 (한글 완벽 지원)"""
    member = get_object_or_404(ChurchMember, id=member_id)
    
    if year is None:
        year = date.today().year
    
    offerings = Offering.objects.filter(
        member=member,
        offering_date__year=year,
        is_confirmed=True
    ).order_by('offering_date')
    
    total = sum([o.amount for o in offerings])
    
    # 증명서 발급 로그 기록
    TaxCertificateIssueLog.objects.create(
        member=member,
        year=year,
        issued_by=request.user,
        issue_type='html'
    )
    
    # HTML 생성
    html = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <title>헌금 증명서</title>
        <style>
            body {{ font-family: 'Malgun Gothic', sans-serif; margin: 40px; }}
            .certificate {{ border: 3px double #000; padding: 30px; max-width: 800px; margin: 0 auto; }}
            .header {{ text-align: center; margin-bottom: 30px; }}
            .header h1 {{ color: #2c3e50; border-bottom: 2px solid #3498db; padding-bottom: 10px; }}
            .info-table {{ width: 100%; border-collapse: collapse; margin: 20px 0; }}
            .info-table th, .info-table td {{ border: 1px solid #ddd; padding: 10px; text-align: left; }}
            .info-table th {{ background: #f8f9fa; }}
            .total {{ font-size: 1.2em; color: #e74c3c; font-weight: bold; margin-top: 20px; }}
            .print-btn {{ margin-top: 30px; padding: 10px 20px; background: #3498db; color: white; border: none; cursor: pointer; }}
        </style>
    </head>
    <body>
        <div class="certificate">
            <div class="header">
                <h1>⛪ 교회 헌금 증명서</h1>
                <h2>{year}년도</h2>
            </div>
            
            <table class="info-table">
                <tr><th>교인번호</th><td>{member.member_id}</td></tr>
                <tr><th>성명</th><td>{member.korean_name}</td></tr>
                <tr><th>영문명</th><td>{member.english_name or '-'}</td></tr>
                <tr><th>생년월일</th><td>{member.birth_date.strftime('%Y년 %m월 %d일') if member.birth_date else '-'}</td></tr>
                <tr><th>세례일</th><td>{member.baptism_date.strftime('%Y년 %m월 %d일') if member.baptism_date else '-'}</td></tr>
                <tr><th>발급일</th><td>{date.today().strftime('%Y년 %m월 %d일')}</td></tr>
            </table>
            
            <h3>헌금 내역</h3>
            <table class="info-table">
                <tr><th>일자</th><th>헌금유형</th><th>금액</th><th>비고</th></tr>
                {"".join([f'<tr><td>{o.offering_date}</td><td>{o.offering_type.name}</td><td>{o.amount:,}원</td><td>{o.notes or "-"}</td></tr>' for o in offerings])}
            </table>
            
            <div class="total">총 합계: {total:,}원</div>
            
            <div style="margin-top: 40px; text-align: center;">
                <button class="print-btn" onclick="printCertificate()">🖨️ 인쇄하기</button>
                <a href="/offerings/certificates/"><button style="margin-left:10px;">← 목록으로</button></a>
            </div>
            
            <div style="margin-top: 50px; border-top: 1px dashed #ccc; padding-top: 20px; font-size: 0.9em; color: #666;">
                <p>※ 본 증명서는 세법 제34조에 의거하여 발급합니다.</p>
                <p>※ 연말정산 시 세무서에 제출하시면 소득공제 혜택을 받으실 수 있습니다.</p>
                <p>※ 교회 인장: <strong>부평우리교회</strong> 직인</p>
            </div>
        </div>
        
        <script>
            // 인쇄 시 로그 기록 및 스타일 조정
            function printCertificate() {{
                // 인쇄 로그 기록
                fetch('/offerings/certificate/{member_id}/{year}/log-print/', {{
                    method: 'POST',
                    headers: {{
                        'X-CSRFToken': getCookie('csrftoken')
                    }}
                }});
                
                // 인쇄 실행
                window.print();
            }}
            
            function getCookie(name) {{
                let cookieValue = null;
                if (document.cookie && document.cookie !== '') {{
                    const cookies = document.cookie.split(';');
                    for (let i = 0; i < cookies.length; i++) {{
                        const cookie = cookies[i].trim();
                        if (cookie.substring(0, name.length + 1) === (name + '=')) {{
                            cookieValue = decodeURIComponent(cookie.substring(name.length + 1));
                            break;
                        }}
                    }}
                }}
                return cookieValue;
            }}
            
            window.onbeforeprint = function() {{
                document.querySelector('.print-btn').style.display = 'none';
            }};
            
            window.onafterprint = function() {{
                document.querySelector('.print-btn').style.display = 'inline-block';
            }};
        </script>
    </body>
    </html>
    """
    
    return HttpResponse(html)

@login_required
def export_offerings_excel(request, year=None):
    """헌금 데이터 엑셀 내보내기"""
    if year is None:
        year = date.today().year
    
    # 데이터 조회
    offerings = Offering.objects.filter(
        offering_date__year=year,
        is_confirmed=True
    ).select_related('member', 'offering_type').order_by('offering_date')
    
    # 워크북 생성
    wb = Workbook()
    
    # 헌금 상세 내역 시트
    ws1 = wb.active
    ws1.title = '헌금상세내역'
    headers = ['헌금일자', '교인번호', '한글이름', '영문이름', '헌금유형', '금액', 
               '결제방법', '은행명', '계좌번호', '참조번호', '확인여부', '비고']
    ws1.append(headers)
    
    # 스타일 설정
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for cell in ws1[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    # 데이터 추가
    type_summary = {}
    member_summary = {}
    monthly_summary = {}
    
    for offering in offerings:
        ws1.append([
            offering.offering_date,
            offering.member.member_id,
            offering.member.korean_name,
            offering.member.english_name,
            offering.offering_type.name,
            offering.amount,
            offering.get_payment_method_display(),
            offering.bank_name or '',
            offering.account_number or '',
            offering.reference_number or '',
            '확인' if offering.is_confirmed else '미확인',
            offering.notes or '',
        ])
        
        # 집계 데이터 수집
        type_name = offering.offering_type.name
        type_summary[type_name] = type_summary.get(type_name, 0) + offering.amount
        
        member_key = (offering.member.member_id, offering.member.korean_name)
        member_summary[member_key] = member_summary.get(member_key, 0) + offering.amount
        
        month = offering.offering_date.month
        monthly_summary[month] = monthly_summary.get(month, 0) + offering.amount
    
    # 유형별 합계 시트
    ws2 = wb.create_sheet('유형별합계')
    ws2.append(['헌금유형', '총액'])
    for cell in ws2[1]:
        cell.fill = header_fill
        cell.font = header_font
    for type_name, total in sorted(type_summary.items()):
        ws2.append([type_name, total])
    
    # 교인별 합계 시트
    ws3 = wb.create_sheet('교인별합계')
    ws3.append(['교인번호', '한글이름', '총액'])
    for cell in ws3[1]:
        cell.fill = header_fill
        cell.font = header_font
    for (member_id, name), total in sorted(member_summary.items()):
        ws3.append([member_id, name, total])
    
    # 월별 합계 시트
    ws4 = wb.create_sheet('월별합계')
    ws4.append(['월', '총액'])
    for cell in ws4[1]:
        cell.fill = header_fill
        cell.font = header_font
    for month in range(1, 13):
        if month in monthly_summary:
            ws4.append([f'{month}월', monthly_summary[month]])
    
    # 응답 생성
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f'헌금내역_{year}년_{datetime.now().strftime("%Y%m%d")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response

@login_required
def export_members_excel(request):
    """교인 데이터 엑셀 내보내기"""
    members = ChurchMember.objects.all()
    
    # 워크북 생성
    wb = Workbook()
    ws = wb.active
    ws.title = '교인명단'
    
    # 헤더
    headers = ['교인번호', '한글이름', '영문이름', '성별', '생년월일', '세례일', 
               '가족번호', '가족관계', '휴대전화', '이메일', '주소', '부서', '직분', 
               '상태', '헌금봉투번호', '세금공제동의', '총헌금액', '등록일']
    ws.append(headers)
    
    # 스타일 설정
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    # 데이터 추가
    for member in members:
        # 해당 교인의 총 헌금액 계산
        total_offering = Offering.objects.filter(
            member=member,
            is_confirmed=True
        ).aggregate(total=models.Sum('amount'))['total'] or 0
        
        ws.append([
            member.member_id,
            member.korean_name,
            member.english_name or '',
            member.get_gender_display(),
            member.birth_date,
            member.baptism_date,
            member.family_id or '',
            member.relationship or '',
            member.phone or '',
            member.email or '',
            member.address or '',
            member.department or '',
            member.position or '',
            member.get_status_display(),
            member.offering_number or '',
            '동의' if member.tax_issuance_consent else '미동의',
            total_offering,
            member.created_at.date(),
        ])
    
    # 응답 생성
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f'교인명단_{datetime.now().strftime("%Y%m%d")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response
    
    return response

@login_required
def export_tax_certificate_excel(request, year=None):
    """연말 증명서용 엑셀 내보내기"""
    if year is None:
        year = date.today().year
    
    members = ChurchMember.objects.filter(tax_issuance_consent=True)
    
    # 워크북 생성
    wb = Workbook()
    ws = wb.active
    ws.title = '증명서용데이터'
    
    # 헤더
    headers = ['교인번호', '성명', '영문명', '주민등록번호', '주소', '연락처', 
               '십일조', '감사헌금', '선교헌금', '건축헌금', '기타헌금', '총헌금액', '발급여부', '비고']
    ws.append(headers)
    
    # 스타일 설정
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    # 통계용 변수
    total_members = 0
    total_amount = 0
    max_amount = 0
    min_amount = float('inf')
    
    for member in members:
        # 해당 연도 헌금 조회
        offerings = Offering.objects.filter(
            member=member,
            offering_date__year=year,
            is_confirmed=True
        )
        
        # 헌금 유형별 합계
        offering_summary = {}
        member_total = 0
        
        for offering in offerings:
            type_name = offering.offering_type.name
            offering_summary[type_name] = offering_summary.get(type_name, 0) + offering.amount
            member_total += offering.amount
        
        if member_total == 0:
            continue
        
        # 십일조, 감사헌금 등 분리
        tithe_amount = offering_summary.get('십일조', 0)
        thanks_amount = offering_summary.get('감사헌금', 0)
        mission_amount = offering_summary.get('선교헌금', 0)
        building_amount = offering_summary.get('건축헌금', 0)
        other_amount = member_total - (tithe_amount + thanks_amount + mission_amount + building_amount)
        
        ws.append([
            member.member_id,
            member.korean_name,
            member.english_name or '',
            '',  # 실제 사용 시 개인정보 보호 주의
            member.address or '',
            member.phone or '',
            tithe_amount,
            thanks_amount,
            mission_amount,
            building_amount,
            other_amount,
            member_total,
            'Y',
            '',
        ])
        
        # 통계 계산
        total_members += 1
        total_amount += member_total
        max_amount = max(max_amount, member_total)
        min_amount = min(min_amount, member_total)
    
    # 통계 시트
    ws2 = wb.create_sheet('통계')
    ws2.append(['항목', '값'])
    for cell in ws2[1]:
        cell.fill = header_fill
        cell.font = header_font
    
    avg_amount = total_amount / total_members if total_members > 0 else 0
    min_amount = min_amount if min_amount != float('inf') else 0
    
    ws2.append(['총 교인수', total_members])
    ws2.append(['총 헌금액', total_amount])
    ws2.append(['평균 헌금액', avg_amount])
    ws2.append(['최대 헌금액', max_amount])
    ws2.append(['최소 헌금액', min_amount])
    
    # 응답 생성
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f'연말증명서용_{year}년_{datetime.now().strftime("%Y%m%d")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response


@login_required
def log_certificate_print(request, member_id, year):
    """증명서 인쇄 로그 기록"""
    if request.method == 'POST':
        member = get_object_or_404(ChurchMember, id=member_id)
        TaxCertificateIssueLog.objects.create(
            member=member,
            year=year,
            issued_by=request.user,
            issue_type='print'
        )
        return HttpResponse('OK')
    return HttpResponse('Method not allowed', status=405)

@login_required
def offering_list(request):
    """헌금 전체 목록 뷰"""
    from .models import Offering
    show = request.GET.get('show', 'all')
    if show == 'unconfirmed':
        offerings = Offering.objects.filter(is_confirmed=False).select_related('member', 'offering_type').order_by('-offering_date', '-id')
    else:
        offerings = Offering.objects.all().select_related('member', 'offering_type').order_by('-offering_date', '-id')
    return render(request, 'offerings/offering_list.html', {'offerings': offerings})