# budget/excel_utils.py
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
from django.utils import timezone

def create_expense_excel(transactions, filters=None):
    """지출 내역 Excel 생성"""
    wb = Workbook()
    ws = wb.active
    ws.title = "지출내역"
    
    # 스타일 정의
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 제목
    ws.merge_cells('A1:J1')
    ws['A1'] = '지출 내역 보고서'
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    # 생성 정보
    ws.merge_cells('A2:J2')
    ws['A2'] = f'생성일시: {timezone.now().strftime("%Y년 %m월 %d일 %H:%M")}'
    ws['A2'].alignment = Alignment(horizontal='right')
    
    # 필터 정보
    if filters:
        ws.merge_cells('A3:J3')
        filter_text = ' | '.join([f"{k}: {v}" for k, v in filters.items() if v])
        ws['A3'] = f'조회 조건: {filter_text}'
        ws['A3'].alignment = Alignment(horizontal='left')
        header_row = 5
    else:
        header_row = 4
    
    # 헤더
    headers = ['번호', '지출일자', '부서', '내용', '거래처', '금액', '상태', '신청자', '결재자', '결재일시']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # 데이터
    row_num = header_row + 1
    total_amount = 0
    
    for idx, transaction in enumerate(transactions, start=1):
        ws.cell(row=row_num, column=1, value=idx).border = border
        ws.cell(row=row_num, column=2, value=transaction.transaction_date.strftime('%Y-%m-%d')).border = border
        ws.cell(row=row_num, column=3, value=transaction.budget.department_name).border = border
        ws.cell(row=row_num, column=4, value=transaction.description).border = border
        ws.cell(row=row_num, column=5, value=transaction.vendor or '-').border = border
        
        amount_cell = ws.cell(row=row_num, column=6, value=int(transaction.amount))
        amount_cell.number_format = '#,##0'
        amount_cell.border = border
        amount_cell.alignment = Alignment(horizontal='right')
        
        status_map = {'pending': '대기', 'approved': '승인', 'rejected': '반려'}
        ws.cell(row=row_num, column=7, value=status_map.get(transaction.status, transaction.status)).border = border
        ws.cell(row=row_num, column=8, value=transaction.requester.username if transaction.requester else '-').border = border
        ws.cell(row=row_num, column=9, value=transaction.approved_by.username if transaction.approved_by else '-').border = border
        ws.cell(row=row_num, column=10, value=transaction.approved_at.strftime('%Y-%m-%d %H:%M') if transaction.approved_at else '-').border = border
        
        if transaction.status == 'approved':
            total_amount += transaction.amount
        
        row_num += 1
    
    # 합계
    ws.cell(row=row_num, column=1, value='합계').font = Font(bold=True)
    ws.cell(row=row_num, column=1).border = border
    ws.merge_cells(f'A{row_num}:E{row_num}')
    ws.cell(row=row_num, column=1).alignment = Alignment(horizontal='center')
    
    total_cell = ws.cell(row=row_num, column=6, value=int(total_amount))
    total_cell.number_format = '#,##0'
    total_cell.font = Font(bold=True)
    total_cell.border = border
    total_cell.alignment = Alignment(horizontal='right')
    total_cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    
    # 열 너비 조정
    column_widths = [8, 12, 15, 30, 15, 15, 10, 12, 12, 18]
    for i, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    # BytesIO로 저장
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return buffer.getvalue()


def create_budget_excel(budgets, year):
    """예산 현황 Excel 생성"""
    wb = Workbook()
    ws = wb.active
    ws.title = f"{year}년 예산"
    
    # 스타일
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 제목
    ws.merge_cells('A1:G1')
    ws['A1'] = f'{year}년 예산 집행 현황'
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    ws.merge_cells('A2:G2')
    ws['A2'] = f'생성일시: {timezone.now().strftime("%Y년 %m월 %d일 %H:%M")}'
    ws['A2'].alignment = Alignment(horizontal='right')
    
    # 헤더
    headers = ['부서명', '예산총액', '집행액', '잔액', '집행률(%)', '승인건수', '대기건수']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # 데이터
    row_num = 5
    total_budget = 0
    total_spent = 0
    
    for budget in budgets:
        spent = budget.spent_amount
        balance = budget.balance
        usage_rate = (spent / budget.total_amount * 100) if budget.total_amount > 0 else 0
        
        ws.cell(row=row_num, column=1, value=budget.department_name).border = border
        
        budget_cell = ws.cell(row=row_num, column=2, value=int(budget.total_amount))
        budget_cell.number_format = '#,##0'
        budget_cell.border = border
        budget_cell.alignment = Alignment(horizontal='right')
        
        spent_cell = ws.cell(row=row_num, column=3, value=int(spent))
        spent_cell.number_format = '#,##0'
        spent_cell.border = border
        spent_cell.alignment = Alignment(horizontal='right')
        
        balance_cell = ws.cell(row=row_num, column=4, value=int(balance))
        balance_cell.number_format = '#,##0'
        balance_cell.border = border
        balance_cell.alignment = Alignment(horizontal='right')
        
        rate_cell = ws.cell(row=row_num, column=5, value=usage_rate)
        rate_cell.number_format = '0.0'
        rate_cell.border = border
        rate_cell.alignment = Alignment(horizontal='center')
        
        ws.cell(row=row_num, column=6, value=budget.transactions.filter(status='approved').count()).border = border
        ws.cell(row=row_num, column=6).alignment = Alignment(horizontal='center')
        
        ws.cell(row=row_num, column=7, value=budget.transactions.filter(status='pending').count()).border = border
        ws.cell(row=row_num, column=7).alignment = Alignment(horizontal='center')
        
        total_budget += budget.total_amount
        total_spent += spent
        
        row_num += 1
    
    # 합계
    ws.cell(row=row_num, column=1, value='합계').font = Font(bold=True)
    ws.cell(row=row_num, column=1).border = border
    
    for col in range(2, 8):
        ws.cell(row=row_num, column=col).border = border
        ws.cell(row=row_num, column=col).fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        ws.cell(row=row_num, column=col).font = Font(bold=True)
    
    ws.cell(row=row_num, column=2, value=int(total_budget)).number_format = '#,##0'
    ws.cell(row=row_num, column=2).alignment = Alignment(horizontal='right')
    
    ws.cell(row=row_num, column=3, value=int(total_spent)).number_format = '#,##0'
    ws.cell(row=row_num, column=3).alignment = Alignment(horizontal='right')
    
    ws.cell(row=row_num, column=4, value=int(total_budget - total_spent)).number_format = '#,##0'
    ws.cell(row=row_num, column=4).alignment = Alignment(horizontal='right')
    
    total_rate = (total_spent / total_budget * 100) if total_budget > 0 else 0
    ws.cell(row=row_num, column=5, value=total_rate).number_format = '0.0'
    ws.cell(row=row_num, column=5).alignment = Alignment(horizontal='center')
    
    # 열 너비
    column_widths = [20, 15, 15, 15, 12, 12, 12]
    for i, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return buffer.getvalue()
